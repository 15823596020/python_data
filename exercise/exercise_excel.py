from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


class ExerciseExcel:
    def __init__(self, excel_title, excel_data):  # 属性有表头信息，数据信息（通过传入的参数得到）
        self.excel_title = excel_title
        self.excel_data = excel_data

    def create_excel(self):  # 定义创建excel，并写入内容
        wt = Workbook()  # 实例化
        sheet = wt.active  # 获取当前页签
        sheet.title = "体检人员信息"  # 设置当前页签名

        # 定义表格的表头
        for i in range(len(excel_title)):
            sheet.cell(row=1, column=i + 1).value = excel_title[i]  # 向第1行中的对应列中写入表头信息
            sheet.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')  # 设置内容居中显示
        # 向表中写入数据
        for i in range(len(excel_data)):  # 循环行
            sheet.cell(row=i + 2, column=1).value = excel_data[i]["姓名"]  # 第1列写入某人的姓名
            sheet.cell(row=i + 2, column=2).value = excel_data[i]["身高"]  # 第2列写入某人的身高
            sheet.cell(row=i + 2, column=3).value = excel_data[i]["体重"]  # 第3列写入某人的体重
            # 设置内容居中显示
            sheet.cell(row=i + 2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=i + 2, column=2).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=i + 2, column=3).alignment = Alignment(horizontal='center', vertical='center')

        wt.save("./exercise_data/examiner_information.xlsx")  # 保存excel文件

    def health_mark(self):  # 定义一个方法来衡量某人的体重是否是健康体
        rt = load_workbook("./exercise_data/examiner_information.xlsx")  # 实例化load_workbook
        sheet_read = rt["体检人员信息"]  # 获取需要操作的页签
        sheet_read['D1'] = "备注"  # 单元格D1中写入备注
        sheet_read.column_dimensions['D'].width = 20  # 调整D这一列的列宽
        # sheet_read.row_dimensions[1].height = 40  # 调整第1行的行高
        sheet_read['D1'].alignment = Alignment(horizontal='center', vertical='center')  # 设置内容居中显示

        for i in range(len(excel_data)):  # 读取某人的相关信息
            name = sheet_read.cell(row=i + 2, column=1).value  # 取出姓名
            height = sheet_read.cell(row=i + 2, column=2).value  # 取出身高
            weight = sheet_read.cell(row=i + 2, column=3).value  # 取出体重

            if weight == (height - 70) * 0.6:  # 如果体重=公式计算出的体重，则为健康体
                print(f"{name}:健康")
                sheet_read.cell(row=i + 2, column=4, value="健康").alignment = Alignment(vertical='center')  # 设置内容垂直居中显示
            elif weight < (height - 70) * 0.6:  # 如果体重<公式计算出的体重，则为偏瘦
                print(f"{name}:偏瘦,健康体重为{(height - 70) * 0.6}")
                sheet_read.cell(row=i + 2, column=4, value=f"偏瘦,健康体重为{(height - 70) * 0.6}").alignment = Alignment(
                    vertical='center')  # 设置内容垂直居中显示
            else:  # 如果体>公式计算出的体重，则为偏胖
                print(f"{name}:偏胖,健康体重为{(height - 70) * 0.6}")
                sheet_read.cell(row=i + 2, column=4, value=f"偏胖,健康体重为{(height - 70) * 0.6}").alignment = Alignment(
                    vertical='center')  # 设置内容垂直居中显示

        rt.save("homework_excel.xlsx")  # 保存excel文件


if __name__ == '__main__':
    # 需要传入的表头列表
    excel_title = ["姓名", "身高", "体重"]
    # 需要传入的数据列表
    excel_data = [{"姓名": "张三", "身高": 170, "体重": 60}, {"姓名": "李四", "身高": 180, "体重": 70},
                  {"姓名": "王五", "身高": 160, "体重": 50}, {"姓名": "史蒂芬", "身高": 175, "体重": 55},
                  {"姓名": "思琪", "身高": 150, "体重": 40}, {"姓名": "小乔", "身高": 165, "体重": 45}]

    exerciseexcel = ExerciseExcel(excel_title, excel_data)  # 实例化，并传入表头信息列表，数据信息列表
    exerciseexcel.create_excel()  # 实例调用create_excel方法
    exerciseexcel.health_mark()  # 实例调用health_mark方法
