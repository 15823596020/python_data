"""
使用openpyxl实现以下需求

使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)
"""
from openpyxl import Workbook, load_workbook


class OpenpyxlExcel:  # 定义一个OpenpyxlExcel类
    def create_excel(self, list_data):  # 定义一个创建文件，并向文件中写入内容的方法
        wt = Workbook()  # 实例化Workbook
        focus = wt.active  # 获取当前活跃的页签
        focus.title = "人员信息"  # 页签取名人员信息
        focus["A1"] = "姓名"  # 单元格A1中写入姓名
        focus["B1"] = "身高"  # 单元格B1中写入身高
        focus["C1"] = "体重"  # 单元格C1中写入体重

        # for循环，将list_data中的数据写入到excel文件中
        for i in range(len(list_data)):  # 循环写的人数
            for j in range(len(list_data[i])):  # 循环写每个人的相关信息
                focus.cell(row=i + 2, column=j + 1).value = list_data[i][j]  # 从第2行开始，第1列写入某人的姓名，第2列写入某人的身高，第3列写入某人的体重

        wt.save("homework_excel.xlsx")  # 保存excel文件

    def health_mark(self, list_data):  # 定义一个方法来衡量某人的体重是否是健康体
        rt = load_workbook(filename="homework_excel.xlsx")  # 实例化load_workbook
        sheet = rt["人员信息"]
        sheet['D1'] = "备注"
        for i in range(len(list_data)):  # 需要读取的某个人的信息
            name = sheet.cell(row=i + 2, column=1).value  # 取出姓名
            height = sheet.cell(row=i + 2, column=2).value  # 取出身高
            weight = sheet.cell(row=i + 2, column=3).value  # 取出体重
            if int(weight) == (int(height) - 70) * 0.6:  # 如果体重=公式计算出的体重，则为健康体
                print(f"{name}:健康")
                sheet.cell(row=i + 2, column=4, value="健康")
            elif int(weight) < (int(height) - 70) * 0.6:  # 如果体重<公式计算出的体重，则为偏瘦
                print(f"{name}:偏瘦,健康体重为{(int(height) - 70) * 0.6}")
                sheet.cell(row=i + 2, column=4, value=f"偏瘦,健康体重为{(int(height) - 70) * 0.6}")
            else:  # 如果体>公式计算出的体重，则为偏胖
                print(f"{name}:偏胖,健康体重为{(int(height) - 70) * 0.6}")
                sheet.cell(row=i + 2, column=4, value=f"偏胖,健康体重为{(int(height) - 70) * 0.6}")

        rt.save("homework_excel.xlsx")  # 保存excel文件


list_data = [("张三", "170", "60"), ("李四", "180", "70"), ("王五", "160", "50"), ("史蒂芬", "175", "55"),
             ("思琪", "150", "40"), ("小乔", "165", "45")]  # 列表包裹着的元素用于存放某人的相关信息

openpyxl_excel = OpenpyxlExcel()  # 实例化OpenpyxlExcel类
openpyxl_excel.create_excel(list_data)  # 实例调用create_excel方法，并传入需要写入的数据列表
openpyxl_excel.health_mark(list_data)  # 实例调用health_mark方法

